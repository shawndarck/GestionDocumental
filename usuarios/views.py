import random
from types import NoneType
from django.conf import settings
from django.core.mail import send_mail
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect
from django.forms import ValidationError
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.contrib.auth import authenticate, login as dj_login, logout
from django.urls import reverse, reverse_lazy
from django.views import generic
from django.views.generic.edit import DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.conf import settings
from django.contrib.auth.models import Group
from usuarios.models import Usuario 
from ciclo_phva.models import (
    ItemEstandar,
    Evidencia,
    Formato,
    SubEstandar,
    Estandar,
    Ciclo,
    Phva,
)

from .forms import (
    EvidenciaModelForm,
    FormatoModelForm,
    EstadoItemForm,
    AccesoUsuarioForm,
    UsuarioForm,
    AdminForm,
    CambiarParsswordForm,
)

from bootstrap_modal_forms.generic import BSModalCreateView
from bootstrap_modal_forms.generic import (
    BSModalLoginView,
    BSModalFormView,
    BSModalCreateView,
    BSModalUpdateView,
    BSModalReadView,
    BSModalDeleteView
)

from verify_email.email_handler import send_verification_email
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

CONTENT_TYPES = ['pdf','png']
# 2.5MB - 2621440
# 5MB - 5242880
# 10MB - 10485760
# 20MB - 20971520
# 50MB - 5242880
# 100MB 104857600
# 250MB - 214958080
# 500MB - 429916160
MAX_UPLOAD_SIZE = "2621440"

class CambiaEstadoUsuario(generic.ListView, LoginRequiredMixin):
    item = Usuario
    context_object_name = 'usuario'
    template_name = 'usuarios/gestion_usuarios.html'
    http_method_names = ['get']

    def get_queryset(self):
        pass

    def post(self, request, *args, **kwargs):
        form=self.form_class(request.POST)
        if form.is_valid():
            messages.success(self.request, "Se a cambiado el estado correctamente")
            return redirect(self.success_url)
        else:
            return HttpResponseRedirect(reverse_lazy('gestion_usuarios'))

    def get_context_data(self, **kwargs):
        # Prerrequisito
        context = super(CambiaEstadoUsuario, self).get_context_data(**kwargs) 
        usu=Usuario.objects.get(id=self.kwargs.get('pk'))
        if usu.is_active == True:
            usu.is_active = False
        else:
            usu.is_active = True
        usu.save(update_fields=['is_active'])
        # Contextos individuales (Objeto)
        messages.success(self.request, "Se a cambiado el estado correctamente")
        context['usuarios'] = Usuario.objects.all()
        return context

    #     success_url = "/planear/"

    # def form_valid(self, form, **kwargs):
    #     self.object = self.get_object()
    #     item = ItemEstandar.objects.get(id=self.object.pk) # Obtener pk de la url con self.object.pk y self.get_object()
    #     puntaje_maximo = item.puntaje_maximo
    #     # Acceder al id de la fk con .id
    #     estado = item.fk_estado.id
    #     if form.instance.fk_estado.id == 1:
    #         item.puntaje_obtenido = puntaje_maximo
    #         item.save(update_fields=['puntaje_obtenido'])
    #     elif form.instance.fk_estado.id == 2 or form.instance.fk_estado.id == 3:
    #         item.puntaje_obtenido = 0
    #         item.save(update_fields=['puntaje_obtenido'])
    #     item.fk_estado = form.instance.fk_estado
    #     item.save(update_fields=['fk_estado_id'])
    #     return HttpResponseRedirect(self.get_success_url())

class GestionUsuarios(generic.ListView, LoginRequiredMixin):
    item = Usuario
    context_object_name = 'usuario'
    template_name = 'usuarios/gestion_usuarios.html'
    success_url = "/gestion_usuarios/"

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        # Prerrequisito
        context = super(GestionUsuarios, self).get_context_data(**kwargs)        
        # Contextos individuales (Objeto)
        context['usuarios'] = Usuario.objects.all()
        return context


class ActuarUsunormal(generic.ListView, LoginRequiredMixin):
    item = ItemEstandar
    context_object_name = 'item_estandar'
    template_name = 'usuarios/usunormal_actuar.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        # Prerrequisito
        context = super(ActuarUsunormal, self).get_context_data(**kwargs)
        # public ItemEstandar lista_items
        
        # Contextos individuales (Objeto)
        context['ciclo'] = Ciclo.objects.get(id = 4)
        context['estandar'] = Estandar.objects.get(id = 7)
        context['sub_estandar'] = SubEstandar.objects.get(id = 21)
        # Items de estandar (Lista)
        context['item_estandar1'] = ItemEstandar.objects.filter(fk_sub_estandar = 21)
        return context


class HacerUsunormal(generic.ListView, LoginRequiredMixin):
    item = ItemEstandar
    context_object_name = 'item_estandar'
    template_name = 'usuarios/usunormal_hacer.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(HacerUsunormal, self).get_context_data(**kwargs)
        
        # Contextos individuales (Objeto)
        context['ciclo'] = Ciclo.objects.get(id = 2)
        context['estandar'] = Estandar.objects.get(id = 3)
        context['estandar2'] = Estandar.objects.get(id = 4)
        context['estandar3'] = Estandar.objects.get(id = 5)
        context['sub_estandar'] = SubEstandar.objects.get(id = 14)
        context['sub_estandar2'] = SubEstandar.objects.get(id = 15)
        context['sub_estandar3'] = SubEstandar.objects.get(id = 16)
        context['sub_estandar4'] = SubEstandar.objects.get(id = 17)
        context['sub_estandar5'] = SubEstandar.objects.get(id = 18)
        context['sub_estandar6'] = SubEstandar.objects.get(id = 19)
        # Items de estandar (Lista)
        context['item_estandar1'] = ItemEstandar.objects.filter(fk_sub_estandar = 14)
        context['item_estandar2'] = ItemEstandar.objects.filter(fk_sub_estandar = 15)
        context['item_estandar3'] = ItemEstandar.objects.filter(fk_sub_estandar = 16)
        context['item_estandar4'] = ItemEstandar.objects.filter(fk_sub_estandar = 17)
        context['item_estandar5'] = ItemEstandar.objects.filter(fk_sub_estandar = 18)
        context['item_estandar6'] = ItemEstandar.objects.filter(fk_sub_estandar = 19)
        return context


class VerificarUsunormal(generic.ListView, LoginRequiredMixin):
    item = ItemEstandar
    context_object_name = 'item_estandar'
    template_name = 'usuarios/usunormal_verificar.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(VerificarUsunormal, self).get_context_data(**kwargs)
        
        # Contextos individuales (Objeto)
        context['ciclo'] = Ciclo.objects.get(id = 3)
        context['estandar'] = Estandar.objects.get(id = 6)
        context['sub_estandar'] = SubEstandar.objects.get(id = 20)
        # Items de estandar (Lista)
        context['item_estandar1'] = ItemEstandar.objects.filter(fk_sub_estandar = 20)
        return context


class PlanearUsunormal(generic.ListView, LoginRequiredMixin):
    item = ItemEstandar
    context_object_name = 'item_estandar'
    template_name = 'usuarios/usunormal_planear.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(PlanearUsunormal, self).get_context_data(**kwargs)
        # Contextos individuales (Objeto)
        context['ciclo'] = Ciclo.objects.get(id = 1)
        context['estandar'] = Estandar.objects.get(id = 1)
        context['estandar2'] = Estandar.objects.get(id = 2)
        context['sub_estandar'] = SubEstandar.objects.get(id = 1)
        context['sub_estandar2'] = SubEstandar.objects.get(id = 2)
        context['sub_estandar3'] = SubEstandar.objects.get(id = 3)
        context['sub_estandar4'] = SubEstandar.objects.get(id = 4)
        context['sub_estandar5'] = SubEstandar.objects.get(id = 5)
        context['sub_estandar6'] = SubEstandar.objects.get(id = 6)
        context['sub_estandar7'] = SubEstandar.objects.get(id = 7)
        context['sub_estandar8'] = SubEstandar.objects.get(id = 8)
        context['sub_estandar9'] = SubEstandar.objects.get(id = 9)
        context['sub_estandar10'] = SubEstandar.objects.get(id = 10)
        context['sub_estandar11'] = SubEstandar.objects.get(id = 11)
        context['sub_estandar12'] = SubEstandar.objects.get(id = 12)
        context['sub_estandar13'] = SubEstandar.objects.get(id = 13)
        # Items de estandar (Lista)
        context['item_estandar1'] = ItemEstandar.objects.filter(fk_sub_estandar = 1)
        context['item_estandar2'] = ItemEstandar.objects.filter(fk_sub_estandar = 2)
        context['item_estandar3'] = ItemEstandar.objects.filter(fk_sub_estandar = 3)
        context['item_estandar4'] = ItemEstandar.objects.filter(fk_sub_estandar = 4)
        context['item_estandar5'] = ItemEstandar.objects.filter(fk_sub_estandar = 5)
        context['item_estandar6'] = ItemEstandar.objects.filter(fk_sub_estandar = 6)
        context['item_estandar7'] = ItemEstandar.objects.filter(fk_sub_estandar = 7)
        context['item_estandar8'] = ItemEstandar.objects.filter(fk_sub_estandar = 8)
        context['item_estandar9'] = ItemEstandar.objects.filter(fk_sub_estandar = 9)
        context['item_estandar10'] = ItemEstandar.objects.filter(fk_sub_estandar = 10)
        context['item_estandar11'] = ItemEstandar.objects.filter(fk_sub_estandar = 11)
        context['item_estandar12'] = ItemEstandar.objects.filter(fk_sub_estandar = 12)
        context['item_estandar13'] = ItemEstandar.objects.filter(fk_sub_estandar = 13)
        return context


class PermisosPlanearCreateView(BSModalCreateView):
    model = ItemEstandar
    template_name = 'usuarios/modal_permisos.html'
    form_class = AccesoUsuarioForm
    success_message = 'Success: Book was updated.'
    success_url = "/accesos_planear/"

    def form_valid(self, form, **kwargs):
        self.object = self.get_object()
        item:(ItemEstandar) = ItemEstandar.objects.get(id=self.object.pk) # Obtener pk de la url con self.object.pk y self.get_object()
        # Acceder al id de la fk con .id
        usu_pk:(int) = self.request.POST.get('users')
        usuario:(Usuario) = Usuario.objects.get(id=usu_pk)
        item.permisos_usuarios.add(usuario)

        return HttpResponseRedirect(self.get_success_url())


class PermisosHacerCreateView(BSModalCreateView):
    model = ItemEstandar
    template_name = 'usuarios/modal_permisos.html'
    form_class = AccesoUsuarioForm
    success_message = 'Success: Book was updated.'
    success_url = "/accesos_hacer/"

    def form_valid(self, form, **kwargs):
        self.object = self.get_object()
        item:(ItemEstandar) = ItemEstandar.objects.get(id=self.object.pk) # Obtener pk de la url con self.object.pk y self.get_object()
        # Acceder al id de la fk con .id
        usu_pk:(int) = self.request.POST.get('users')
        usuario:(Usuario) = Usuario.objects.get(id=usu_pk)
        item.permisos_usuarios.add(usuario)

        return HttpResponseRedirect(self.get_success_url())


class PermisosVerificarCreateView(BSModalCreateView):
    model = ItemEstandar
    template_name = 'usuarios/modal_permisos.html'
    form_class = AccesoUsuarioForm
    success_message = 'Success: Book was updated.'
    success_url = "/accesos_verificar/"

    def form_valid(self, form, **kwargs):
        self.object = self.get_object()
        item:(ItemEstandar) = ItemEstandar.objects.get(id=self.object.pk) # Obtener pk de la url con self.object.pk y self.get_object()
        # Acceder al id de la fk con .id
        usu_pk:(int) = self.request.POST.get('users')
        usuario:(Usuario) = Usuario.objects.get(id=usu_pk)
        item.permisos_usuarios.add(usuario)

        return HttpResponseRedirect(self.get_success_url())


class PermisosActuarCreateView(BSModalCreateView):
    model = ItemEstandar
    template_name = 'usuarios/modal_permisos.html'
    form_class = AccesoUsuarioForm
    success_message = 'Success: Book was updated.'
    success_url = "/accesos_actuar/"

    def form_valid(self, form, **kwargs):
        self.object = self.get_object()
        item:(ItemEstandar) = ItemEstandar.objects.get(id=self.object.pk) # Obtener pk de la url con self.object.pk y self.get_object()
        # Acceder al id de la fk con .id
        usu_pk:(int) = self.request.POST.get('users')
        usuario:(Usuario) = Usuario.objects.get(id=usu_pk)
        item.permisos_usuarios.add(usuario)

        return HttpResponseRedirect(self.get_success_url())


class PermisosReadView(generic.ListView):
    model = Usuario
    template_name = 'usuarios/leer_accesos.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(PermisosReadView, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk')
        context['usuarios'] =  Usuario.objects.filter(itemestandar__in=[pk])
        context['item'] = ItemEstandar.objects.get(id=pk)
        return context



def quitar_acceso(request, pk, id):
    # dictionary for initial data with  
    # field names as keys 
    context ={} 

    # fetch the object related to passed id 
    usu = get_object_or_404(Usuario, id = pk)
    item = get_object_or_404(ItemEstandar, id = id)


    if request.method =="POST": 
        # delete object 
        usu.delete() 
        # after deleting redirect to  
        # home page 
        return HttpResponseRedirect("/") 

    return reverse_lazy('planear')

class QuitarAcceso(generic.ListView, LoginRequiredMixin):
    item = Usuario
    context_object_name = 'usuario'
    template_name = 'usuarios/eliminar_acceso.html'
    http_method_names = ['post', 'get']

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        # Prerrequisito
        context = super(QuitarAcceso, self).get_context_data(**kwargs) 
        context['usu'] = Usuario.objects.get(id=self.kwargs.get('pk'))
        return context


    def post(self, request, *args, **kwargs):
        usu = Usuario.objects.get(id=self.kwargs.get('pk'))
        item = ItemEstandar.objects.get(id=self.kwargs.get('id'))
        item.permisos_usuarios.remove(usu)
        return HttpResponseRedirect(reverse_lazy('accesos_planear'))




class GestionPermisosEvidencias(generic.ListView, LoginRequiredMixin):
    item = Usuario
    context_object_name = 'item'
    template_name = 'usuarios/evidencias_usuarios.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(GestionPermisosEvidencias, self).get_context_data(**kwargs)
        context['usuarios'] = Usuario.objects.filter(es_usuario = True)
        return context


class AccesosPlanear(generic.ListView, LoginRequiredMixin):
    item = ItemEstandar
    context_object_name = 'item'
    template_name = 'usuarios/accesos_planear.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(AccesosPlanear, self).get_context_data(**kwargs)
        # Contextos individuales (Objeto)
        context['ciclo'] = Ciclo.objects.get(id = 1)
        context['estandar'] = Estandar.objects.get(id = 1)
        context['estandar2'] = Estandar.objects.get(id = 2)
        context['sub_estandar'] = SubEstandar.objects.get(id = 1)
        context['sub_estandar2'] = SubEstandar.objects.get(id = 2)
        context['sub_estandar3'] = SubEstandar.objects.get(id = 3)
        context['sub_estandar4'] = SubEstandar.objects.get(id = 4)
        context['sub_estandar5'] = SubEstandar.objects.get(id = 5)
        context['sub_estandar6'] = SubEstandar.objects.get(id = 6)
        context['sub_estandar7'] = SubEstandar.objects.get(id = 7)
        context['sub_estandar8'] = SubEstandar.objects.get(id = 8)
        context['sub_estandar9'] = SubEstandar.objects.get(id = 9)
        context['sub_estandar10'] = SubEstandar.objects.get(id = 10)
        context['sub_estandar11'] = SubEstandar.objects.get(id = 11)
        context['sub_estandar12'] = SubEstandar.objects.get(id = 12)
        context['sub_estandar13'] = SubEstandar.objects.get(id = 13)
        # Items de estandar (Lista)
        context['item_estandar1'] = ItemEstandar.objects.filter(fk_sub_estandar = 1)
        context['item_estandar2'] = ItemEstandar.objects.filter(fk_sub_estandar = 2)
        context['item_estandar3'] = ItemEstandar.objects.filter(fk_sub_estandar = 3)
        context['item_estandar4'] = ItemEstandar.objects.filter(fk_sub_estandar = 4)
        context['item_estandar5'] = ItemEstandar.objects.filter(fk_sub_estandar = 5)
        context['item_estandar6'] = ItemEstandar.objects.filter(fk_sub_estandar = 6)
        context['item_estandar7'] = ItemEstandar.objects.filter(fk_sub_estandar = 7)
        context['item_estandar8'] = ItemEstandar.objects.filter(fk_sub_estandar = 8)
        context['item_estandar9'] = ItemEstandar.objects.filter(fk_sub_estandar = 9)
        context['item_estandar10'] = ItemEstandar.objects.filter(fk_sub_estandar = 10)
        context['item_estandar11'] = ItemEstandar.objects.filter(fk_sub_estandar = 11)
        context['item_estandar12'] = ItemEstandar.objects.filter(fk_sub_estandar = 12)
        context['item_estandar13'] = ItemEstandar.objects.filter(fk_sub_estandar = 13)
        return context


class AccesosHacer(generic.ListView, LoginRequiredMixin):
    item = ItemEstandar
    context_object_name = 'item'
    template_name = 'usuarios/accesos_hacer.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(AccesosHacer, self).get_context_data(**kwargs)
        # Contextos individuales (Objeto)
        context['ciclo'] = Ciclo.objects.get(id = 2)
        context['estandar'] = Estandar.objects.get(id = 3)
        context['estandar2'] = Estandar.objects.get(id = 4)
        context['estandar3'] = Estandar.objects.get(id = 5)
        context['sub_estandar'] = SubEstandar.objects.get(id = 14)
        context['sub_estandar2'] = SubEstandar.objects.get(id = 15)
        context['sub_estandar3'] = SubEstandar.objects.get(id = 16)
        context['sub_estandar4'] = SubEstandar.objects.get(id = 17)
        context['sub_estandar5'] = SubEstandar.objects.get(id = 18)
        context['sub_estandar6'] = SubEstandar.objects.get(id = 19)
        # Items de estandar (Lista)
        context['item_estandar1'] = ItemEstandar.objects.filter(fk_sub_estandar = 14)
        context['item_estandar2'] = ItemEstandar.objects.filter(fk_sub_estandar = 15)
        context['item_estandar3'] = ItemEstandar.objects.filter(fk_sub_estandar = 16)
        context['item_estandar4'] = ItemEstandar.objects.filter(fk_sub_estandar = 17)
        context['item_estandar5'] = ItemEstandar.objects.filter(fk_sub_estandar = 18)
        context['item_estandar6'] = ItemEstandar.objects.filter(fk_sub_estandar = 19)
        return context


class AccesosVerificar(generic.ListView, LoginRequiredMixin):
    item = ItemEstandar
    context_object_name = 'item'
    template_name = 'usuarios/accesos_verificar.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(AccesosVerificar, self).get_context_data(**kwargs)
        # Contextos individuales (Objeto)
        context['ciclo'] = Ciclo.objects.get(id = 3)
        context['estandar'] = Estandar.objects.get(id = 6)
        context['sub_estandar'] = SubEstandar.objects.get(id = 20)
        # Items de estandar (Lista)
        context['item_estandar1'] = ItemEstandar.objects.filter(fk_sub_estandar = 20)
        return context


class AccesosActuar(generic.ListView, LoginRequiredMixin):
    item = ItemEstandar
    context_object_name = 'item'
    template_name = 'usuarios/accesos_actuar.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(AccesosActuar, self).get_context_data(**kwargs)

        # Contextos individuales (Objeto)
        context['ciclo'] = Ciclo.objects.get(id = 4)
        context['estandar'] = Estandar.objects.get(id = 7)
        context['sub_estandar'] = SubEstandar.objects.get(id = 21)
        # Items de estandar (Lista)
        context['item_estandar1'] = ItemEstandar.objects.filter(fk_sub_estandar = 21)
        return context


class EvidenciaPlanearCreateView(BSModalCreateView):
    model = ItemEstandar
    template_name = 'usuarios/crear_evidencia.html'
    form_class = EvidenciaModelForm
    success_message = 'Success: Book was created.'
    success_url = reverse_lazy('planear')

    def form_valid(self, form):
        form.instance.fk_item_estandar_id = self.kwargs.get('pk')
        return super(EvidenciaPlanearCreateView, self).form_valid(form)

class EvidenciaHacerCreateView(BSModalCreateView):
    model = ItemEstandar
    template_name = 'usuarios/crear_evidencia.html'
    form_class = EvidenciaModelForm
    success_message = 'Success: Book was created.'
    success_url = reverse_lazy('hacer')

    def form_valid(self, form):
        form.instance.fk_item_estandar_id = self.kwargs.get('pk')
        return super(EvidenciaHacerCreateView, self).form_valid(form)

class EvidenciaVerificarCreateView(BSModalCreateView):
    model = ItemEstandar
    template_name = 'usuarios/crear_evidencia.html'
    form_class = EvidenciaModelForm
    success_message = 'Success: Book was created.'
    success_url = reverse_lazy('verificar')

    def form_valid(self, form):
        form.instance.fk_item_estandar_id = self.kwargs.get('pk')
        return super(EvidenciaVerificarCreateView, self).form_valid(form)

class EvidenciaActuarCreateView(BSModalCreateView):
    model = ItemEstandar
    template_name = 'usuarios/crear_evidencia.html'
    form_class = EvidenciaModelForm
    success_message = 'Success: Book was created.'
    success_url = reverse_lazy('actuar')

    def form_valid(self, form):
        form.instance.fk_item_estandar_id = self.kwargs.get('pk')
        return super(EvidenciaActuarCreateView, self).form_valid(form)


class EvidenciaUsuarioReadView(generic.ListView):
    model = Evidencia
    template_name = 'usuarios/leer_evidencias.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(EvidenciaUsuarioReadView, self).get_context_data(**kwargs)
        pk = self.kwargs.get('fk')
        items = ItemEstandar.objects.filter(permisos_usuarios__in=[self.request.user.id])
        item:(ItemEstandar) = ItemEstandar
        for i in items:
            if i.id == pk:
                item = i
                context['evidencias'] = Evidencia.objects.filter(fk_item_estandar = item.id)
                break
        return context


class EvidenciaReadView(generic.ListView):
    model = Evidencia
    template_name = 'usuarios/leer_evidencias.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(EvidenciaReadView, self).get_context_data(**kwargs)
        pk = self.kwargs.get('fk')
        context['evidencias'] = Evidencia.objects.filter(fk_item_estandar = pk)
        return context


class Planear(generic.ListView, LoginRequiredMixin):
    item = ItemEstandar
    context_object_name = 'item_estandar'
    template_name = 'usuarios/planear.html'
    

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(Planear, self).get_context_data(**kwargs)
        lista_items:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=1)
        lista_items2:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=2)
        lista_items3:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=3)
        lista_items4:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=4)
        lista_items5:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=5)
        lista_items6:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=6)
        lista_items7:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=7)
        lista_items8:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=8)
        lista_items9:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=9)
        lista_items10:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=10)
        lista_items11:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=11)
        lista_items12:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=12)
        lista_items13:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=13)
        
        lista_calculo_estandar:(SubEstandar) = SubEstandar.objects.filter(fk_estandar_id=1)
        lista_calculo_estandar2:(SubEstandar) = SubEstandar.objects.filter(fk_estandar_id=2)

        lista_estandares:(Estandar) = Estandar.objects.filter(fk_ciclo_id=1)

        #Campos para acceder a la bd (Actualizar)
        sub_estandar:(SubEstandar) = SubEstandar.objects.get(id=1)
        sub_estandar2:(SubEstandar) = SubEstandar.objects.get(id=2)
        sub_estandar3:(SubEstandar) = SubEstandar.objects.get(id=3)
        sub_estandar4:(SubEstandar) = SubEstandar.objects.get(id=4)
        sub_estandar5:(SubEstandar) = SubEstandar.objects.get(id=5)
        sub_estandar6:(SubEstandar) = SubEstandar.objects.get(id=6)
        sub_estandar7:(SubEstandar) = SubEstandar.objects.get(id=7)
        sub_estandar8:(SubEstandar) = SubEstandar.objects.get(id=8)
        sub_estandar9:(SubEstandar) = SubEstandar.objects.get(id=9)
        sub_estandar10:(SubEstandar) = SubEstandar.objects.get(id=10)
        sub_estandar11:(SubEstandar) = SubEstandar.objects.get(id=11)
        sub_estandar12:(SubEstandar) = SubEstandar.objects.get(id=12)
        sub_estandar13:(SubEstandar) = SubEstandar.objects.get(id=13)
        
        estandar:(Estandar) = Estandar.objects.get(id=1)
        estandar2:(Estandar) = Estandar.objects.get(id=2)
        
        ciclo:(Ciclo) = Ciclo.objects.get(id=1)
        ciclos:(Ciclo) = Ciclo.objects.all()
        phva:(Phva) = Phva.objects.get(id=1)
        acumulador:(float) = 0
        coefiente:(float) = 0
        #Terminar el otro sub estandar (lista y for)
        for i in lista_items:
            acumulador += i.puntaje_obtenido
        sub_estandar.calificacion_obtenida = acumulador
        if sub_estandar.calificacion_obtenida != 0:
            coefiente = sub_estandar.calificacion_maxima / sub_estandar.calificacion_obtenida
            sub_estandar.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar.porcentaje_obtenido = 0
        sub_estandar.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_items2:
            acumulador += i.puntaje_obtenido
        sub_estandar2.calificacion_obtenida = acumulador
        if sub_estandar2.calificacion_obtenida != 0:
            coefiente = sub_estandar2.calificacion_maxima / sub_estandar2.calificacion_obtenida
            sub_estandar2.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar2.porcentaje_obtenido = 0
        sub_estandar2.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_items3:
            acumulador += i.puntaje_obtenido
        sub_estandar3.calificacion_obtenida = acumulador
        if sub_estandar3.calificacion_obtenida != 0:
            coefiente = sub_estandar3.calificacion_maxima / sub_estandar3.calificacion_obtenida
            sub_estandar3.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar3.porcentaje_obtenido = 0
        sub_estandar3.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_items4:
            acumulador += i.puntaje_obtenido
        sub_estandar4.calificacion_obtenida = acumulador
        if sub_estandar4.calificacion_obtenida != 0:
            coefiente = sub_estandar4.calificacion_maxima / sub_estandar4.calificacion_obtenida
            sub_estandar4.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar4.porcentaje_obtenido = 0
        sub_estandar4.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_items5:
            acumulador += i.puntaje_obtenido
        sub_estandar5.calificacion_obtenida = acumulador
        if sub_estandar5.calificacion_obtenida != 0:
            coefiente = sub_estandar5.calificacion_maxima / sub_estandar5.calificacion_obtenida
            sub_estandar5.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar5.porcentaje_obtenido = 0
        sub_estandar5.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_items6:
            acumulador += i.puntaje_obtenido
        sub_estandar6.calificacion_obtenida = acumulador
        if sub_estandar6.calificacion_obtenida != 0:
            coefiente = sub_estandar6.calificacion_maxima / sub_estandar6.calificacion_obtenida
            sub_estandar6.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar6.porcentaje_obtenido = 0
        sub_estandar6.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_items7:
            acumulador += i.puntaje_obtenido
        sub_estandar7.calificacion_obtenida = acumulador
        if sub_estandar7.calificacion_obtenida != 0:
            coefiente = sub_estandar7.calificacion_maxima / sub_estandar7.calificacion_obtenida
            sub_estandar7.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar7.porcentaje_obtenido = 0
        sub_estandar7.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_items8:
            acumulador += i.puntaje_obtenido
        sub_estandar8.calificacion_obtenida = acumulador
        if sub_estandar8.calificacion_obtenida != 0:
            coefiente = sub_estandar8.calificacion_maxima / sub_estandar8.calificacion_obtenida
            sub_estandar8.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar8.porcentaje_obtenido = 0
        sub_estandar8.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_items9:
            acumulador += i.puntaje_obtenido
        sub_estandar9.calificacion_obtenida = acumulador
        if sub_estandar9.calificacion_obtenida != 0:
            coefiente = sub_estandar9.calificacion_maxima / sub_estandar9.calificacion_obtenida
            sub_estandar9.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar9.porcentaje_obtenido = 0
        sub_estandar9.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_items10:
            acumulador += i.puntaje_obtenido
        sub_estandar10.calificacion_obtenida = acumulador
        if sub_estandar10.calificacion_obtenida != 0:
            coefiente = sub_estandar10.calificacion_maxima / sub_estandar10.calificacion_obtenida
            sub_estandar10.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar10.porcentaje_obtenido = 0
        sub_estandar10.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_items11:
            acumulador += i.puntaje_obtenido
        sub_estandar11.calificacion_obtenida = acumulador
        if sub_estandar11.calificacion_obtenida != 0:
            coefiente = sub_estandar11.calificacion_maxima / sub_estandar11.calificacion_obtenida
            sub_estandar11.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar11.porcentaje_obtenido = 0
        sub_estandar11.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_items12:
            acumulador += i.puntaje_obtenido
        sub_estandar12.calificacion_obtenida = acumulador
        if sub_estandar12.calificacion_obtenida != 0:
            coefiente = sub_estandar12.calificacion_maxima / sub_estandar12.calificacion_obtenida
            sub_estandar12.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar12.porcentaje_obtenido = 0
        sub_estandar12.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_items13:
            acumulador += i.puntaje_obtenido
        sub_estandar13.calificacion_obtenida = acumulador
        if sub_estandar13.calificacion_obtenida != 0:
            coefiente = sub_estandar13.calificacion_maxima / sub_estandar13.calificacion_obtenida
            sub_estandar13.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar13.porcentaje_obtenido = 0
        sub_estandar13.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_calculo_estandar:
            acumulador += i.calificacion_obtenida
        estandar.calificacion_obtenida = acumulador
        if estandar.calificacion_obtenida != 0:
            coefiente = estandar.calificacion_maxima / estandar.calificacion_obtenida
            estandar.porcentaje_obtenido = 100 / coefiente
        else:
            estandar.porcentaje_obtenido = 0
        estandar.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_calculo_estandar2:
            acumulador += i.calificacion_obtenida
        estandar2.calificacion_obtenida = acumulador
        if estandar2.calificacion_obtenida != 0:
            coefiente = estandar2.calificacion_maxima / estandar2.calificacion_obtenida
            estandar2.porcentaje_obtenido = 100 / coefiente
        else:
            estandar2.porcentaje_obtenido = 0
        estandar2.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_estandares:
            acumulador += i.calificacion_obtenida
        ciclo.calificacion_obtenida = acumulador
        if ciclo.calificacion_obtenida != 0:
            coefiente = ciclo.calificacion_maxima / ciclo.calificacion_obtenida
            ciclo.porcentaje_obtenido = 100 / coefiente
        else:
            ciclo.porcentaje_obtenido = 0
        ciclo.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in ciclos:
            acumulador += i.calificacion_obtenida
        phva.calificacion_obtenida = acumulador
        phva.save(update_fields=['calificacion_obtenida'])

        # Contextos individuales (Objeto)
        context['ciclo'] = Ciclo.objects.get(id = 1)
        context['estandar'] = Estandar.objects.get(id = 1)
        context['estandar2'] = Estandar.objects.get(id = 2)
        context['sub_estandar'] = SubEstandar.objects.get(id = 1)
        context['sub_estandar2'] = SubEstandar.objects.get(id = 2)
        context['sub_estandar3'] = SubEstandar.objects.get(id = 3)
        context['sub_estandar4'] = SubEstandar.objects.get(id = 4)
        context['sub_estandar5'] = SubEstandar.objects.get(id = 5)
        context['sub_estandar6'] = SubEstandar.objects.get(id = 6)
        context['sub_estandar7'] = SubEstandar.objects.get(id = 7)
        context['sub_estandar8'] = SubEstandar.objects.get(id = 8)
        context['sub_estandar9'] = SubEstandar.objects.get(id = 9)
        context['sub_estandar10'] = SubEstandar.objects.get(id = 10)
        context['sub_estandar11'] = SubEstandar.objects.get(id = 11)
        context['sub_estandar12'] = SubEstandar.objects.get(id = 12)
        context['sub_estandar13'] = SubEstandar.objects.get(id = 13)
        # Items de estandar (Lista)
        context['item_estandar1'] = ItemEstandar.objects.filter(fk_sub_estandar = 1)
        context['item_estandar2'] = ItemEstandar.objects.filter(fk_sub_estandar = 2)
        context['item_estandar3'] = ItemEstandar.objects.filter(fk_sub_estandar = 3)
        context['item_estandar4'] = ItemEstandar.objects.filter(fk_sub_estandar = 4)
        context['item_estandar5'] = ItemEstandar.objects.filter(fk_sub_estandar = 5)
        context['item_estandar6'] = ItemEstandar.objects.filter(fk_sub_estandar = 6)
        context['item_estandar7'] = ItemEstandar.objects.filter(fk_sub_estandar = 7)
        context['item_estandar8'] = ItemEstandar.objects.filter(fk_sub_estandar = 8)
        context['item_estandar9'] = ItemEstandar.objects.filter(fk_sub_estandar = 9)
        context['item_estandar10'] = ItemEstandar.objects.filter(fk_sub_estandar = 10)
        context['item_estandar11'] = ItemEstandar.objects.filter(fk_sub_estandar = 11)
        context['item_estandar12'] = ItemEstandar.objects.filter(fk_sub_estandar = 12)
        context['item_estandar13'] = ItemEstandar.objects.filter(fk_sub_estandar = 13)
        return context


class Hacer(generic.ListView, LoginRequiredMixin):
    item = ItemEstandar
    context_object_name = 'item_estandar'
    template_name = 'usuarios/hacer.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(Hacer, self).get_context_data(**kwargs)
        lista_items:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=14)
        lista_items2:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=15)
        lista_items3:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=16)
        lista_items4:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=17)
        lista_items5:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=18)
        lista_items6:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=19)
        
        lista_calculo_estandar:(SubEstandar) = SubEstandar.objects.filter(fk_estandar_id=3)
        lista_calculo_estandar2:(SubEstandar) = SubEstandar.objects.filter(fk_estandar_id=4)
        lista_calculo_estandar3:(SubEstandar) = SubEstandar.objects.filter(fk_estandar_id=5)

        lista_estandares:(Estandar) = Estandar.objects.filter(fk_ciclo_id=2)

        #Campos para acceder a la bd (Actualizar)
        sub_estandar:(SubEstandar) = SubEstandar.objects.get(id=14)
        sub_estandar2:(SubEstandar) = SubEstandar.objects.get(id=15)
        sub_estandar3:(SubEstandar) = SubEstandar.objects.get(id=16)
        sub_estandar4:(SubEstandar) = SubEstandar.objects.get(id=17)
        sub_estandar5:(SubEstandar) = SubEstandar.objects.get(id=18)
        sub_estandar6:(SubEstandar) = SubEstandar.objects.get(id=19)
        
        estandar:(Estandar) = Estandar.objects.get(id=3)
        estandar2:(Estandar) = Estandar.objects.get(id=4)
        estandar3:(Estandar) = Estandar.objects.get(id=5)

        ciclos:(Ciclo) = Ciclo.objects.all()
        phva:(Phva) = Phva.objects.get(id=1)
        
        ciclo:(Ciclo) = Ciclo.objects.get(id=2)
        acumulador:(int) = 0
        coefiente:(float) = 0
        #Terminar el otro sub estandar (lista y for)

        for i in lista_items:
            acumulador += i.puntaje_obtenido
        sub_estandar.calificacion_obtenida = acumulador
        if sub_estandar.calificacion_obtenida != 0:
            coefiente = sub_estandar.calificacion_maxima / sub_estandar.calificacion_obtenida
            sub_estandar.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar.porcentaje_obtenido = 0
        sub_estandar.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0

        for i in lista_items2:
            acumulador += i.puntaje_obtenido
        sub_estandar2.calificacion_obtenida = acumulador
        if sub_estandar2.calificacion_obtenida != 0:
            coefiente = sub_estandar2.calificacion_maxima / sub_estandar2.calificacion_obtenida
            sub_estandar2.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar2.porcentaje_obtenido = 0
        sub_estandar2.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0

        for i in lista_items3:
            acumulador += i.puntaje_obtenido
        sub_estandar3.calificacion_obtenida = acumulador
        if sub_estandar3.calificacion_obtenida != 0:
            coefiente = sub_estandar3.calificacion_maxima / sub_estandar3.calificacion_obtenida
            sub_estandar3.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar3.porcentaje_obtenido = 0
        sub_estandar3.save(update_fields=[ 'calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0

        for i in lista_items4:
            acumulador += i.puntaje_obtenido
        sub_estandar4.calificacion_obtenida = acumulador
        if sub_estandar4.calificacion_obtenida != 0:
            coefiente = sub_estandar4.calificacion_maxima / sub_estandar4.calificacion_obtenida
            sub_estandar4.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar4.porcentaje_obtenido = 0
        sub_estandar4.save(update_fields=[ 'calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0

        for i in lista_items5:
            acumulador += i.puntaje_obtenido
        sub_estandar5.calificacion_obtenida = acumulador
        if sub_estandar5.calificacion_obtenida != 0:
            coefiente = sub_estandar5.calificacion_maxima / sub_estandar5.calificacion_obtenida
            sub_estandar5.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar5.porcentaje_obtenido = 0
        sub_estandar5.save(update_fields=[ 'calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0

        for i in lista_items6:
            acumulador += i.puntaje_obtenido
        sub_estandar6.calificacion_obtenida = acumulador
        if sub_estandar6.calificacion_obtenida != 0:
            coefiente = sub_estandar6.calificacion_maxima / sub_estandar6.calificacion_obtenida
            sub_estandar6.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar6.porcentaje_obtenido = 0
        sub_estandar6.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0

        for i in lista_calculo_estandar:
            acumulador += i.calificacion_obtenida
        estandar.calificacion_obtenida = acumulador
        if estandar.calificacion_obtenida != 0:
            coefiente = estandar.calificacion_maxima / estandar.calificacion_obtenida
            estandar.porcentaje_obtenido = 100 / coefiente
        else:
            estandar.porcentaje_obtenido = 0
        estandar.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0

        for i in lista_calculo_estandar2:
            acumulador += i.calificacion_obtenida
        estandar2.calificacion_obtenida = acumulador
        if estandar2.calificacion_obtenida != 0:
            coefiente = estandar2.calificacion_maxima / estandar2.calificacion_obtenida
            estandar2.porcentaje_obtenido = 100 / coefiente
        else:
            estandar2.porcentaje_obtenido = 0
        estandar2.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0

        for i in lista_calculo_estandar3:
            acumulador += i.calificacion_obtenida
        estandar3.calificacion_obtenida = acumulador
        if estandar3.calificacion_obtenida != 0:
            coefiente = estandar3.calificacion_maxima / estandar3.calificacion_obtenida
            estandar3.porcentaje_obtenido = 100 / coefiente
        else:
            estandar3.porcentaje_obtenido = 0
        estandar3.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0

        for i in lista_estandares:
            acumulador += i.calificacion_obtenida
        ciclo.calificacion_obtenida = acumulador
        if ciclo.calificacion_obtenida != 0:
            coefiente = ciclo.calificacion_maxima / ciclo.calificacion_obtenida
            ciclo.porcentaje_obtenido = 100 / coefiente
        else:
            ciclo.porcentaje_obtenido = 0
        ciclo.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0

        for i in ciclos:
            acumulador += i.calificacion_obtenida
        phva.calificacion_obtenida = acumulador
        phva.save(update_fields=['calificacion_obtenida'])

        # Contextos individuales (Objeto)
        context['ciclo'] = Ciclo.objects.get(id = 2)
        context['estandar'] = Estandar.objects.get(id = 3)
        context['estandar2'] = Estandar.objects.get(id = 4)
        context['estandar3'] = Estandar.objects.get(id = 5)
        context['sub_estandar'] = SubEstandar.objects.get(id = 14)
        context['sub_estandar2'] = SubEstandar.objects.get(id = 15)
        context['sub_estandar3'] = SubEstandar.objects.get(id = 16)
        context['sub_estandar4'] = SubEstandar.objects.get(id = 17)
        context['sub_estandar5'] = SubEstandar.objects.get(id = 18)
        context['sub_estandar6'] = SubEstandar.objects.get(id = 19)
        # Items de estandar (Lista)
        context['item_estandar1'] = ItemEstandar.objects.filter(fk_sub_estandar = 14)
        context['item_estandar2'] = ItemEstandar.objects.filter(fk_sub_estandar = 15)
        context['item_estandar3'] = ItemEstandar.objects.filter(fk_sub_estandar = 16)
        context['item_estandar4'] = ItemEstandar.objects.filter(fk_sub_estandar = 17)
        context['item_estandar5'] = ItemEstandar.objects.filter(fk_sub_estandar = 18)
        context['item_estandar6'] = ItemEstandar.objects.filter(fk_sub_estandar = 19)
        return context


class Verificar(generic.ListView, LoginRequiredMixin):
    item = ItemEstandar
    context_object_name = 'item_estandar'
    template_name = 'usuarios/verificar.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(Verificar, self).get_context_data(**kwargs)
        lista_items:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=20)

        lista_calculo_estandar:(SubEstandar) = SubEstandar.objects.filter(fk_estandar_id=6)

        lista_estandares:(Estandar) = Estandar.objects.filter(fk_ciclo_id=3)

        #Campos para acceder a la bd (Actualizar)
        sub_estandar:(SubEstandar) = SubEstandar.objects.get(id=20)
        
        estandar:(Estandar) = Estandar.objects.get(id=6)

        ciclos:(Ciclo) = Ciclo.objects.all()
        phva:(Phva) = Phva.objects.get(id=1)
        
        ciclo:(Ciclo) = Ciclo.objects.get(id=3)
        acumulador:(int) = 0
        coefiente:(float) = 0
        #Terminar el otro sub estandar (lista y for)
        for i in lista_items:
            acumulador += i.puntaje_obtenido
        sub_estandar.calificacion_obtenida = acumulador
        if sub_estandar.calificacion_obtenida != 0:
            coefiente = sub_estandar.calificacion_maxima / sub_estandar.calificacion_obtenida
            sub_estandar.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar.porcentaje_obtenido = 0
        sub_estandar.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_calculo_estandar:
            acumulador += i.calificacion_obtenida
        estandar.calificacion_obtenida = acumulador
        if estandar.calificacion_obtenida != 0:
            coefiente = estandar.calificacion_maxima / estandar.calificacion_obtenida
            estandar.porcentaje_obtenido = 100 / coefiente
        else:
            estandar.porcentaje_obtenido = 0
        estandar.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_estandares:
            acumulador += i.calificacion_obtenida
        ciclo.calificacion_obtenida = acumulador
        if ciclo.calificacion_obtenida != 0:
            coefiente = ciclo.calificacion_maxima / ciclo.calificacion_obtenida
            ciclo.porcentaje_obtenido = 100 / coefiente
        else:
            ciclo.porcentaje_obtenido = 0
        ciclo.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in ciclos:
            acumulador += i.calificacion_obtenida
        phva.calificacion_obtenida = acumulador
        phva.save(update_fields=['calificacion_obtenida'])

        # Contextos individuales (Objeto)
        context['ciclo'] = Ciclo.objects.get(id = 3)
        context['estandar'] = Estandar.objects.get(id = 6)
        context['sub_estandar'] = SubEstandar.objects.get(id = 20)
        # Items de estandar (Lista)
        context['item_estandar1'] = ItemEstandar.objects.filter(fk_sub_estandar = 20)
        return context


class Actuar(generic.ListView, LoginRequiredMixin):
    item = ItemEstandar
    context_object_name = 'item_estandar'
    template_name = 'usuarios/actuar.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        # Prerrequisito
        context = super(Actuar, self).get_context_data(**kwargs)
        # public ItemEstandar lista_items
        lista_items:(ItemEstandar) = ItemEstandar.objects.filter(fk_sub_estandar=21)

        lista_calculo_estandar:(SubEstandar) = SubEstandar.objects.filter(fk_estandar_id=7)

        lista_estandares:(Estandar) = Estandar.objects.filter(fk_ciclo_id=4)

        #Campos para acceder a la bd (Actualizar)
        sub_estandar:(SubEstandar) = SubEstandar.objects.get(id=21)
        
        estandar:(Estandar) = Estandar.objects.get(id=7)

        ciclos:(Ciclo) = Ciclo.objects.all()
        phva:(Phva) = Phva.objects.get(id=1)
        
        ciclo:(Ciclo) = Ciclo.objects.get(id=4)
        acumulador:(int) = 0
        coefiente:(float) = 0
        #Terminar el otro sub estandar (lista y for)
        for i in lista_items:
            acumulador += i.puntaje_obtenido
        sub_estandar.calificacion_obtenida = acumulador
        if sub_estandar.calificacion_obtenida != 0:
            coefiente = sub_estandar.calificacion_maxima / sub_estandar.calificacion_obtenida
            sub_estandar.porcentaje_obtenido = 100 / coefiente
        else:
            sub_estandar.porcentaje_obtenido = 0
        sub_estandar.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_calculo_estandar:
            acumulador += i.calificacion_obtenida
        estandar.calificacion_obtenida = acumulador
        if estandar.calificacion_obtenida != 0:
            coefiente = estandar.calificacion_maxima / estandar.calificacion_obtenida
            estandar.porcentaje_obtenido = 100 / coefiente
        else:
            estandar.porcentaje_obtenido = 0
        estandar.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in lista_estandares:
            acumulador += i.calificacion_obtenida
        ciclo.calificacion_obtenida = acumulador
        if ciclo.calificacion_obtenida != 0:
            coefiente = ciclo.calificacion_maxima / ciclo.calificacion_obtenida
            ciclo.porcentaje_obtenido = 100 / coefiente
        else:
            ciclo.porcentaje_obtenido = 0
        ciclo.save(update_fields=['calificacion_obtenida', 'porcentaje_obtenido'])
        acumulador = 0
        coefiente = 0

        for i in ciclos:
            acumulador += i.calificacion_obtenida
        phva.calificacion_obtenida = acumulador
        phva.save(update_fields=['calificacion_obtenida'])

        # Contextos individuales (Objeto)
        context['ciclo'] = Ciclo.objects.get(id = 4)
        context['estandar'] = Estandar.objects.get(id = 7)
        context['sub_estandar'] = SubEstandar.objects.get(id = 21)
        # Items de estandar (Lista)
        context['item_estandar1'] = ItemEstandar.objects.filter(fk_sub_estandar = 21)
        return context


class FormatoCreateView(BSModalCreateView):
    template_name = 'usuarios/crear_formato.html'
    model = Formato
    form_class = FormatoModelForm
    success_message = 'Success: Book was created.'
    success_url = "/formatos_admin/"


class FormatosAdmin(generic.ListView, LoginRequiredMixin):
    item = Formato
    template_name = 'usuarios/formatos.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(FormatosAdmin, self).get_context_data(**kwargs)
        context['formatos'] = Formato.objects.all()
        return context


class FormatosUsunormal(generic.ListView, LoginRequiredMixin):
    item = Formato
    template_name = 'usuarios/usuario/formatos.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(FormatosUsunormal, self).get_context_data(**kwargs)
        context['formatos'] = Formato.objects.all()
        return context


class FormatoDeleteView(BSModalDeleteView):
    model = Formato
    template_name = 'usuarios/eliminar_formato.html'
    success_message = 'Success: Formato borrado.'
    success_url = "/formatos_admin/"


class EvidenciaPlanearDeleteView(BSModalDeleteView):
    model = Evidencia
    template_name = 'usuarios/eliminar_evidencia.html'
    success_message = 'Success: evidencia borrado.'
    success_url = "/planear/"


class ItemEstadoUpdateView(BSModalUpdateView):
    model = ItemEstandar
    template_name = 'usuarios/cambiar_estado_item.html'
    form_class = EstadoItemForm
    success_message = 'Success: Book was updated.'
    success_url = "/planear/"

    def form_valid(self, form, **kwargs):
        self.object = self.get_object()
        item = ItemEstandar.objects.get(id=self.object.pk) # Obtener pk de la url con self.object.pk y self.get_object()
        puntaje_maximo = item.puntaje_maximo
        # Acceder al id de la fk con .id
        estado = item.fk_estado.id
        if form.instance.fk_estado.id == 1:
            item.puntaje_obtenido = puntaje_maximo
            item.save(update_fields=['puntaje_obtenido'])
        elif form.instance.fk_estado.id == 2 or form.instance.fk_estado.id == 3:
            item.puntaje_obtenido = 0
            item.save(update_fields=['puntaje_obtenido'])
        item.fk_estado = form.instance.fk_estado
        item.save(update_fields=['fk_estado_id'])
        return HttpResponseRedirect(self.get_success_url())


class ItemEstadoHacerUpdateView(BSModalUpdateView):
    model = ItemEstandar
    template_name = 'usuarios/cambiar_estado_item.html'
    form_class = EstadoItemForm
    success_message = 'Success: Book was updated.'
    success_url = "/hacer/"

    def form_valid(self, form, **kwargs):
        self.object = self.get_object()
        item = ItemEstandar.objects.get(id=self.object.pk) # Obtener pk de la url con self.object.pk y self.get_object()
        puntaje_maximo = item.puntaje_maximo
        # Acceder al id de la fk con .id
        estado = item.fk_estado.id
        if form.instance.fk_estado.id == 1:
            item.puntaje_obtenido = puntaje_maximo
            item.save(update_fields=['puntaje_obtenido'])
        elif form.instance.fk_estado.id == 2 or form.instance.fk_estado.id == 3:
            item.puntaje_obtenido = 0
            item.save(update_fields=['puntaje_obtenido'])
        item.fk_estado = form.instance.fk_estado
        item.save(update_fields=['fk_estado_id'])
        return HttpResponseRedirect(self.get_success_url())


class ItemEstadoVerificarUpdateView(BSModalUpdateView):
    model = ItemEstandar
    template_name = 'usuarios/cambiar_estado_item.html'
    form_class = EstadoItemForm
    success_message = 'Success: Book was updated.'
    success_url = "/verificar/"

    def form_valid(self, form, **kwargs):
        self.object = self.get_object()
        item = ItemEstandar.objects.get(id=self.object.pk) # Obtener pk de la url con self.object.pk y self.get_object()
        puntaje_maximo = item.puntaje_maximo
        # Acceder al id de la fk con .id
        estado = item.fk_estado.id
        if form.instance.fk_estado.id == 1:
            item.puntaje_obtenido = puntaje_maximo
            item.save(update_fields=['puntaje_obtenido'])
        elif form.instance.fk_estado.id == 2 or form.instance.fk_estado.id == 3:
            item.puntaje_obtenido = 0
            item.save(update_fields=['puntaje_obtenido'])
        item.fk_estado = form.instance.fk_estado
        item.save(update_fields=['fk_estado_id'])
        return HttpResponseRedirect(self.get_success_url())


class ItemEstadoActuarUpdateView(BSModalUpdateView):
    model = ItemEstandar
    template_name = 'usuarios/cambiar_estado_item.html'
    form_class = EstadoItemForm
    success_message = 'Success: Book was updated.'
    success_url = "/actuar/"

    def form_valid(self, form, **kwargs):
        self.object = self.get_object()
        item:(ItemEstandar) = ItemEstandar.objects.get(id=self.object.pk) # Obtener pk de la url con self.object.pk y self.get_object()
        puntaje_maximo = item.puntaje_maximo
        # Acceder al id de la fk con .id

        if form.instance.fk_estado.id == 1:
            item.puntaje_obtenido = puntaje_maximo
            item.save(update_fields=['puntaje_obtenido'])
        elif form.instance.fk_estado.id == 2 or form.instance.fk_estado.id == 3:
            item.puntaje_obtenido = 0
            item.save(update_fields=['puntaje_obtenido'])
        item.fk_estado = form.instance.fk_estado
        item.save(update_fields=['fk_estado_id'])
        return HttpResponseRedirect(self.get_success_url())



def verificar(request):
    return render(request, 'usuarios/verificar.html')


def actuar(request):
    return render(request, 'usuarios/actuar.html')

class PerfilDetailView(generic.ListView, LoginRequiredMixin):
    item = Usuario
    context_object_name = 'usuario'
    template_name = 'usuarios/perfil.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(PerfilDetailView, self).get_context_data(**kwargs)
        context['usuario'] = Usuario.objects.get(id=self.request.user.id)


class ExcelDetailView(generic.ListView, LoginRequiredMixin):
    item = Usuario
    context_object_name = 'usuario'
    template_name = 'usuarios/tortal_administrador.html'

    def get(self, request, *args, **kwargs):
        objeto_a_trabajar = ItemEstandar.objects.all()
        phva = Phva.objects.get(id=1)
        # Abrir flujo de edición de datos
        wb = Workbook()
        ws = wb.active
        contador:(int) = 2
        # Encabezados
        #ws["A" + '1'].alignment = Alignment(horizontal='left')
        ws['A' + '1'] = 'Ciclo'
        ws.alignment = Alignment(horizontal='left')   
        ws['B' + '1'] = 'Estandar'
        ws['D' + '1'] = 'Item de estandar'
        ws['E' + '1'] = 'Valor maximo'
        ws['F' + '1'] = 'Peso porcentual'
        ws['G' + '1'] = 'Puntaje obtenido'
        # Ciclos
        ws['A' + '2'] = 'Planear'
        ws.merge_cells('A2:A23')
        ws['A' + '24'] = 'Hacer'
        ws.merge_cells('A24:A53')
        ws['A' + '54'] = 'Verificar'
        ws.merge_cells('A54:A57')
        ws['A' + '58'] = 'Actuar'
        ws.merge_cells('A58:A61')
        #Estandares
        ws['B' + '2'] = 'Recursos'
        ws.merge_cells('B2:B12')
        ws['B' + '13'] = 'Gestion integral del sistema de gestion de la seguridad y salud en el trabajo'
        ws.merge_cells('B13:B23')
        ws['B' + '24'] = 'Gestion de la salud'
        ws.merge_cells('B24:B41')
        ws['B' + '24'] = 'Gestion de peligros y riesgos'
        ws.merge_cells('B42:B51')
        ws['B' + '52'] = 'Gestion de amenazas'
        ws.merge_cells('B52:B53')
        ws['B' + '54'] = 'Verificacion del SG-SST'
        ws.merge_cells('B54:B57')
        ws['B' + '58'] = 'Mejoramiento'
        ws.merge_cells('B58:B61')
        # Sub estandares
        ws['C' + '2'] = 'Recursos financieros, técnicos humanos y de otra índole requeridos para coordinar y desarrollar el Sistema de Gestion de la Seguridad y Salud en el Trabajo (SG-SST)'
        ws.merge_cells('C2:C9')
        ws['C' + '10'] = 'Capacitación en el Sistema de Gestión de Seguridad y Salud en el Trabajo'
        ws.merge_cells('C10:C12')
        ws['C' + '13'] = 'Política de Seguridad y Salud en el Trabajo'
        ws['C' + '14'] = 'Objetivos del Sistema de Gestión de la Seguridad y la Salud en el Trabajo SG-SST'
        ws['C' + '15'] = 'Evaluación inicial del SG-SST '
        ws['C' + '16'] = 'Plan Anual de Trabajo'
        ws['C' + '17'] = 'Conservación de la documentación'
        ws['C' + '18'] = 'Rendición de cuentas'
        ws['C' + '19'] = 'Normatividad nacional vigente y aplicable en materia de seguridad y salud en el trabajo'
        ws['C' + '20'] = 'Comunicación'
        ws['C' + '21'] = 'Adquisiciones'
        ws['C' + '22'] = 'Contratación'
        ws['C' + '23'] = 'Gestión del cambio'
        ws['C' + '24'] = 'Condiciones de salud en el trabajo'
        ws.merge_cells('C24:C32')
        ws['C' + '33'] = 'Registro, reporte e investigación de las enfermedades laborales, los incidentes y accidentes del trabajo'
        ws.merge_cells('C33:C35')
        ws['C' + '36'] = 'Mecanismos de vigilancia de las condiciones de salud de los trabajadores '
        ws.merge_cells('C36:C41')
        ws['C' + '42'] = 'Identificación de peligros, evaluación y valoración de riesgos'
        ws.merge_cells('C42:C45')
        ws['C' + '46'] = 'Medidas de prevención y control para intervenir los peligros/riesgos'
        ws.merge_cells('C46:C51')
        ws['C' + '52'] = 'Plan de prevención, preparación y respuesta ante emergencias'
        ws.merge_cells('C52:C53')
        ws['C' + '54'] = 'Gestión y resultados del SG-SST'
        ws.merge_cells('C54:C57')
        ws['C' + '58'] = 'Acciones preventivas y correctivas con base en los resultados del SG-SST'
        ws.merge_cells('C58:C61')
        # Peso porcentual
        ws['F' + '2'] = '4'
        ws.merge_cells('F2:F9')
        ws['F' + '10'] = '6'
        ws.merge_cells('F10:F12')
        ws['F' + '13'] = '15'
        ws.merge_cells('F13:F23')
        ws['F' + '24'] = '9'
        ws.merge_cells('F24:F32')
        ws['F' + '33'] = '5'
        ws.merge_cells('F33:F35')
        ws['F' + '36'] = '6'
        ws.merge_cells('F36:F41')
        ws['F' + '42'] = '15'
        ws.merge_cells('F42:F45')
        ws['F' + '46'] = '15'
        ws.merge_cells('F46:F51')
        ws['F' + '52'] = '10'
        ws.merge_cells('F52:F53')
        ws['F' + '54'] = '5'
        ws.merge_cells('F54:F57')
        ws['F' + '58'] = '10'
        ws.merge_cells('F58:F61')
        # Item de estandar
        for q in objeto_a_trabajar:
            column = str(contador) 
            ws['D' + column] = q.descripcion
            ws['E'+ column] = q.puntaje_maximo
            ws['G'+ column] = q.puntaje_obtenido
            contador += 1

        # Calculo de phva
        column = str(contador) 
        ws['F'+ column] = 'Total'
        ws['G'+ column] = phva.calificacion_obtenida
        
        # ws.merge_cells('A1:A9')
        ws.merge_cells('B1:C1')
        nombre_archivo = "ListadoPhva.xlsx"
        # Response HTTP
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        # Ciclo de escritura terminado
        wb.save(response)
        # Retornar el response
        return response

# class PermisosReadView(generic.ListView):
#     model = Usuario
#     template_name = 'usuarios/leer_accesos.html'

#     def get_queryset(self):
#         pass

#     def get_context_data(self, **kwargs):
#         context = super(PermisosReadView, self).get_context_data(**kwargs)
#         pk = self.kwargs.get('pk')
#         context['usuarios'] =  Usuario.objects.filter(itemestandar__in=[pk])
#         return context

def bar_chart_phva(request):
    labels = []
    data = []

    queryset = Ciclo.objects.order_by('-id')[:4]
    for ciclo in queryset:
        labels.append(ciclo.descripcion)
        data.append(ciclo.calificacion_obtenida)

    return render(request, 'usuarios/grafico.html', {
        'labels': labels,
        'data': data,
    })

class PasswordUpdateView(BSModalUpdateView):
    model = Usuario
    template_name = 'usuarios/actualizar_clave.html'
    form_class = CambiarParsswordForm
    success_message = 'Success: Clave cambiada.'
    success_url = reverse_lazy('perfil')

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, {'form': self.form_class})

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            user = Usuario.objects.filter(id=request.user.id)
            if user.exists():
                user = user.first()
                user.set_password(form.cleaned_data.get('password1'))
                user.save()
                logout(request)
                return redirect(self.success_url)
            return redirect(self.success_url)
        else:
            form = self.form_class(request.POST)
            return render(request, self.template_name, {'form': form})


class RegistrarAdministrador(BSModalCreateView):
    template_name = 'usuarios/crear_administrador.html'
    form_class = AdminForm
    success_message = 'Success: Usuario creado.'
    success_url = reverse_lazy('gestion_usuarios')
    
    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            form = self.get_form()
            usu = form.save(commit=False)
            # Enviar correo con clave
            email = form.cleaned_data['email']
            username = form.cleaned_data['username']
            password = form.cleaned_data['password1']
            password = Usuario.objects.make_random_password()
            usu.set_password(password)
            usu.save()
            # Envía mail
            send_mail(
                subject=username,
                message='Has sido registrado en el sistema de gestión documental SST, tu clave es: ' + password,
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[email],
            )
            return self.form_valid(form)
        else:
            return HttpResponseRedirect(reverse_lazy('gestion_usuarios'))


class RegistrarUsuario(BSModalCreateView):
    template_name = 'usuarios/crear_usuario.html'
    form_class = UsuarioForm
    success_message = 'Success: Usuario creado.'
    success_url = reverse_lazy('gestion_usuarios')

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if form.is_valid():
            form = self.get_form()
            usu = form.save(commit=False)
            # Enviar correo con clave
            email = form.cleaned_data['email']
            username = form.cleaned_data['username']
            password = form.cleaned_data['password1']
            password = Usuario.objects.make_random_password()
            usu.set_password(password)

            usu.save()

            send_mail(
                subject=username,
                message='Has sido registrado en el sistema de gestión documental SST, tu clave es: ' + password,
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[email],
            )
            return self.form_valid(form)
        else:
            return HttpResponseRedirect(reverse_lazy('gestion_usuarios'))

def login(request):

    if request.method == 'GET':
        context = ''
        return render(request, 'usuarios/login.html', {'context': context})

    elif request.method == 'POST':
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            dj_login(request, user)
            # Agregar otro if para verificar el estado y validacion de usuario
            # if user.is_active & user.es_valido:
            # Validar grupos (roles)
            if user.groups.filter(name='perfilnormal').exists() & user.es_valido & user.es_usuario:
                return redirect('/torta_usunormal/')
            elif user.groups.filter(name='perfiladministrador').exists() & user.es_valido & user.es_administrador:
                return redirect('/torta_administrador/')
            elif user.groups.filter(name='perfilgestor').exists() & user.es_valido & user.es_gestor:
                return redirect('/torta_gestor/')
            # Si el usuario no esta validado y no tiene grupo se le asigna un grupo y se activa
            elif user.es_usuario == True:
                Group.objects.get(name='perfilnormal').user_set.add(user)
                return redirect('/torta_usunormal/')
            elif user.es_administrador == True:
                Group.objects.get(name='perfiladministrador').user_set.add(user)
                return redirect('/torta_administrador/')
            else:
                context = {'error': 'Wrong credintials'}  # Agregar mensaje de error
                return render(request, 'usuarios/login.html', {'context': context})
        else:
            context = {'error': 'Wrong credintials'}  # Agregar mensaje de error
            return render(request, 'usuarios/login.html', {'context': context})


def calificar_planear(request):
    item_estandar = request.GET.get('cumple', None)
    response = {
        'respuesta': ItemEstandar.objects.filter(puntaje=item_estandar).exists()
    }
    return JsonResponse(response)


class TortaAdministrador(generic.ListView, LoginRequiredMixin):
    item = Ciclo
    context_object_name = 'item'
    template_name = 'usuarios/torta_administrador.html'

    def get_queryset(self):
        pass

    def get_context_data(self, **kwargs):
        context = super(TortaAdministrador, self).get_context_data(**kwargs)
        context['planear'] = Ciclo.objects.get(id = 1)
        context['hacer'] = Ciclo.objects.get(id = 2)
        context['verificar'] = Ciclo.objects.get(id = 3)
        context['actuar'] = Ciclo.objects.get(id = 4)
        context['phva'] = Phva.objects.get(id = 1)
        return context


@login_required
def torta_gestor(request):
    return render(request, 'usuarios/torta_gestor.html')


@login_required
def torta_usunormal(request):
    return render(request, 'usuarios/torta_usunormal.html')
