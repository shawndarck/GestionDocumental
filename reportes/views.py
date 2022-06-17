import datetime
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.views import generic

from usuarios.models import Usuario

from ciclo_phva.models import (
    ItemEstandar,
    Phva,
)
from django.contrib.auth.mixins import LoginRequiredMixin

# Importaciones de libreria para generar reportes
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class ExcelDetailView(generic.ListView, LoginRequiredMixin):
    item = Usuario
    context_object_name = 'usuario'
    template_name = 'usuarios/tortal_administrador.html'

    def get(self, request, *args, **kwargs):
        item_estandar:(ItemEstandar) = ItemEstandar.objects.all()
        phva:(Phva) = Phva.objects.get(id=1)
        fecha = str(datetime.date.today())
        # Abrir flujo de edición de datos
        wb = Workbook()
        ws = wb.active
        contador:(int) = 3
        # Encabezados
        # Título
        ws['A' + '1'] = 'Reporte de ciclo phva'
        ws.merge_cells('A1:E1')
        #Negrilla
        ws['A1'].font = Font(bold=True)
        #Fecha
        ws['F' + '1'] = 'fecha ' + fecha
        ws.merge_cells('F1:G1')
        #Negrilla
        ws['F1'].font = Font(bold=True)
        #ws["A" + '1'].alignment = Alignment(horizontal='left')
        ws['A' + '2'] = 'Ciclo'
        #Negrilla
        ws['A2'].font = Font(bold=True)   
        ws['B' + '2'] = 'Estandar'
        #Negrilla
        ws['B2'].font = Font(bold=True)
        ws['D' + '2'] = 'Item de estandar'
        #Negrilla
        ws['D2'].font = Font(bold=True)
        ws['E' + '2'] = 'Valor maximo'
        #Negrilla
        ws['E2'].font = Font(bold=True)
        ws['F' + '2'] = 'Peso porcentual'
        #Negrilla
        ws['F2'].font = Font(bold=True)
        ws['G' + '2'] = 'Puntaje obtenido'
        #Negrilla
        ws['G2'].font = Font(bold=True)
        # Ciclos
        ws['A' + '3'] = 'Planear'
        ws.merge_cells('A3:A24')
        ws['A' + '25'] = 'Hacer'
        ws.merge_cells('A25:A54')
        ws['A' + '55'] = 'Verificar'
        ws.merge_cells('A55:A58')
        ws['A' + '59'] = 'Actuar'
        ws.merge_cells('A59:A62')
        #Estandares
        ws['B' + '3'] = 'Recursos'
        ws.merge_cells('B3:B13')
        ws['B' + '14'] = 'Gestion integral del sistema de gestion de la seguridad y salud en el trabajo'
        ws.merge_cells('B14:B24')
        ws['B' + '25'] = 'Gestion de la salud'
        ws.merge_cells('B25:B42')
        ws['B' + '43'] = 'Gestion de peligros y riesgos'
        ws.merge_cells('B43:B52')
        ws['B' + '53'] = 'Gestion de amenazas'
        ws.merge_cells('B53:B54')
        ws['B' + '55'] = 'Verificacion del SG-SST'
        ws.merge_cells('B55:B58')
        ws['B' + '59'] = 'Mejoramiento'
        ws.merge_cells('B59:B62')
        # Sub estandares
        ws['C' + '3'] = 'Recursos financieros, técnicos humanos y de otra índole requeridos para coordinar y desarrollar el Sistema de Gestion de la Seguridad y Salud en el Trabajo (SG-SST)'
        ws.merge_cells('C3:C10')
        ws['C' + '11'] = 'Capacitación en el Sistema de Gestión de Seguridad y Salud en el Trabajo'
        ws.merge_cells('C11:C13')
        ws['C' + '14'] = 'Política de Seguridad y Salud en el Trabajo'
        ws['C' + '15'] = 'Objetivos del Sistema de Gestión de la Seguridad y la Salud en el Trabajo SG-SST'
        ws['C' + '16'] = 'Evaluación inicial del SG-SST '
        ws['C' + '17'] = 'Plan Anual de Trabajo'
        ws['C' + '18'] = 'Conservación de la documentación'
        ws['C' + '19'] = 'Rendición de cuentas'
        ws['C' + '20'] = 'Normatividad nacional vigente y aplicable en materia de seguridad y salud en el trabajo'
        ws['C' + '21'] = 'Comunicación'
        ws['C' + '22'] = 'Adquisiciones'
        ws['C' + '23'] = 'Contratación'
        ws['C' + '24'] = 'Gestión del cambio'
        ws['C' + '25'] = 'Condiciones de salud en el trabajo'
        ws.merge_cells('C25:C33')
        ws['C' + '34'] = 'Registro, reporte e investigación de las enfermedades laborales, los incidentes y accidentes del trabajo'
        ws.merge_cells('C34:C36')
        ws['C' + '37'] = 'Mecanismos de vigilancia de las condiciones de salud de los trabajadores '
        ws.merge_cells('C37:C42')
        ws['C' + '43'] = 'Identificación de peligros, evaluación y valoración de riesgos'
        ws.merge_cells('C43:C46')
        ws['C' + '47'] = 'Medidas de prevención y control para intervenir los peligros/riesgos'
        ws.merge_cells('C47:C52')
        ws['C' + '53'] = 'Plan de prevención, preparación y respuesta ante emergencias'
        ws.merge_cells('C53:C54')
        ws['C' + '55'] = 'Gestión y resultados del SG-SST'
        ws.merge_cells('C55:C58')
        ws['C' + '59'] = 'Acciones preventivas y correctivas con base en los resultados del SG-SST'
        ws.merge_cells('C59:C62')
        # Peso porcentual
        ws['F' + '3'] = '4'
        ws.merge_cells('F3:F10')
        ws['F' + '11'] = '6'
        ws.merge_cells('F11:F13')
        ws['F' + '14'] = '15'
        ws.merge_cells('F14:F24')
        ws['F' + '25'] = '9'
        ws.merge_cells('F25:F33')
        ws['F' + '34'] = '5'
        ws.merge_cells('F34:F36')
        ws['F' + '37'] = '6'
        ws.merge_cells('F37:F42')
        ws['F' + '43'] = '15'
        ws.merge_cells('F43:F46')
        ws['F' + '47'] = '15'
        ws.merge_cells('F47:F52')
        ws['F' + '53'] = '10'
        ws.merge_cells('F53:F54')
        ws['F' + '55'] = '5'
        ws.merge_cells('F55:F58')
        ws['F' + '59'] = '10'
        ws.merge_cells('F59:F62')
        # Item de estandar
        for q in item_estandar:
            column = str(contador) 
            ws['D' + column] = q.descripcion
            ws['E'+ column] = q.puntaje_maximo
            ws['G'+ column] = q.puntaje_obtenido
            contador += 1

        # Calculo de phva
        column = str(contador) 
        ws['F'+ column] = 'Total'
        ws['G'+ column] = phva.calificacion_obtenida
        
        # ws.merge_cells('A1:A9')
        ws.merge_cells('B2:C2')
        nombre_archivo:(str) = "ListadoPhva.xlsx"
        # Response HTTP
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        # Ciclo de escritura terminado
        wb.save(response)
        # Retornar el response
        return response
